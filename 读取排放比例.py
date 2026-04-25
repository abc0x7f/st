from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook


PROVINCE_MAP = {
    "Beijing": "北京",
    "Tianjin": "天津",
    "Hebei": "河北",
    "Shanxi": "山西",
    "InnerMongolia": "内蒙古",
    "Liaoning": "辽宁",
    "Jilin": "吉林",
    "Heilongjiang": "黑龙江",
    "Shanghai": "上海",
    "Jiangsu": "江苏",
    "Zhejiang": "浙江",
    "Anhui": "安徽",
    "Fujian": "福建",
    "Jiangxi": "江西",
    "Shandong": "山东",
    "Henan": "河南",
    "Hubei": "湖北",
    "Hunan": "湖南",
    "Guangdong": "广东",
    "Guangxi": "广西",
    "Hainan": "海南",
    "Chongqing": "重庆",
    "Sichuan": "四川",
    "Guizhou": "贵州",
    "Yunnan": "云南",
    "Shaanxi": "陕西",
    "Gansu": "甘肃",
    "Qinghai": "青海",
    "Ningxia": "宁夏",
    "Xinjiang": "新疆",
}


COAL_RELATED_COLUMNS = [
    "Raw_Coal",
    "CleanedCoal",
    "Other_Washed_Coal",
    "Briquettes",
    "Coke",
    "Coke_Oven_Gas",
    "Other_Gas",
    "Other_Coking_Products",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="批量读取排放清单，计算煤炭相关排放占总排放比重（es 的近似代理变量）。"
    )
    parser.add_argument(
        "--input-dir",
        default=r"data\2015-2022各省排放清单(2)",
        help="排放清单所在目录，默认读取 data\\2015-2022各省排放清单(2)",
    )
    parser.add_argument(
        "--output",
        default=r"data\排放比例_es近似.csv",
        help="输出文件名，支持 .csv 或 .xlsx",
    )
    return parser.parse_args()


def normalize_province(sheet_name: str) -> str:
    base = re.sub(r"\d{4}$", "", sheet_name).strip()
    return PROVINCE_MAP.get(base, base)


def parse_year(file_path: Path, sheet_name: str) -> int | None:
    text = f"{file_path.stem} {sheet_name}"
    match = re.search(r"(20\d{2})", text)
    return int(match.group(1)) if match else None


def get_total_emissions_row(ws) -> dict[str, float]:
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header = [str(x).strip() if x is not None else "" for x in header]

    for row in rows:
        if row and row[0] == "TotalEmissions":
            record: dict[str, float] = {}
            for col_name, value in zip(header[1:], row[1:]):
                if not col_name:
                    continue
                record[col_name] = 0.0 if value is None else float(value)
            return record

    raise ValueError(f"工作表 {ws.title} 未找到 TotalEmissions 行")


def extract_one_workbook(file_path: Path) -> list[dict]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    records: list[dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name.upper() == "NOTE":
            continue

        ws = wb[sheet_name]
        total_row = get_total_emissions_row(ws)
        total_emissions = total_row.get("Scope_1_Total", 0.0)
        if total_emissions == 0:
            total_emissions = sum(total_row.values())

        coal_related_emissions = sum(total_row.get(col, 0.0) for col in COAL_RELATED_COLUMNS)
        es_proxy = coal_related_emissions / total_emissions if total_emissions else None

        records.append(
            {
                "year": parse_year(file_path, sheet_name),
                "province": normalize_province(sheet_name),
                "total_emissions_mt_co2": total_emissions,
                "coal_related_emissions_mt_co2": coal_related_emissions,
                "es_proxy": es_proxy,
                "source_file": file_path.name,
                "source_sheet": sheet_name,
            }
        )

    return records


def sort_records(records: list[dict]) -> list[dict]:
    return sorted(records, key=lambda x: (x.get("year") or 0, x.get("province") or ""))


def save_records(records: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "year",
        "province",
        "total_emissions_mt_co2",
        "coal_related_emissions_mt_co2",
        "es_proxy",
        "source_file",
        "source_sheet",
    ]

    if output_path.suffix.lower() == ".xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        ws.append(headers)
        for record in records:
            ws.append([record.get(h) for h in headers])
        wb.save(output_path)
        return

    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(records)


def main() -> None:
    args = parse_args()
    input_dir = Path(args.input_dir)
    output_path = Path(args.output)

    workbook_paths = sorted(
        p for p in input_dir.glob("*.xlsx") if not p.name.startswith("~$")
    )
    if not workbook_paths:
        raise FileNotFoundError(f"未在目录中找到 xlsx 文件: {input_dir}")

    all_records: list[dict] = []
    for file_path in workbook_paths:
        all_records.extend(extract_one_workbook(file_path))

    all_records = sort_records(all_records)
    save_records(all_records, output_path)

    print(f"已输出 {len(all_records)} 行到: {output_path}")
    print("说明：es_proxy 使用煤炭相关排放 / 总排放进行近似，不等同于标准能源结构指标。")


if __name__ == "__main__":
    main()
