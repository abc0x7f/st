from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook

#
# python connect.py --left process2.csv --right 排放比例_es近似.csv --output prcd/process2_es_proxy.csv --include-right es_proxy --how left
#

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="将两个规范化的 xlsx/csv 文件按 year、province 连接，左表支持排除列，右表使用白名单保留列。"
    )
    parser.add_argument("--left", required=True, help="左表文件名，支持 .csv/.xlsx")
    parser.add_argument("--right", required=True, help="右表文件名，支持 .csv/.xlsx")
    parser.add_argument("--output", required=True, help="输出文件名，支持 .csv/.xlsx")
    parser.add_argument(
        "--on",
        nargs="+",
        default=["year", "province"],
        help="连接键，默认 year province",
    )
    parser.add_argument(
        "--exclude-left",
        nargs="*",
        default=[],
        help="左表中连接前要排除的列名",
    )
    parser.add_argument(
        "--include-right",
        nargs="*",
        default=[],
        help="右表中连接前要保留的列名；留空则默认保留全部列",
    )
    parser.add_argument(
        "--how",
        default="inner",
        choices=["inner", "left", "right", "outer"],
        help="连接方式，默认 inner",
    )
    parser.add_argument(
        "--suffixes",
        nargs=2,
        default=["_left", "_right"],
        help="重名列后缀，默认 _left _right",
    )
    return parser.parse_args()


def read_csv(path: Path) -> list[dict]:
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def read_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    records: list[dict] = []
    for row in rows[1:]:
        records.append({header: value for header, value in zip(headers, row)})
    return records


def read_table(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix in {".xlsx", ".xls"}:
        return read_xlsx(path)
    raise ValueError(f"不支持的文件类型: {path}")


def save_csv(records: list[dict], headers: list[str], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(records)


def save_xlsx(records: list[dict], headers: list[str], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(headers)
    for record in records:
        ws.append([record.get(h) for h in headers])
    wb.save(path)


def save_table(records: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    headers: list[str] = []
    for record in records:
        for key in record.keys():
            if key not in headers:
                headers.append(key)

    suffix = path.suffix.lower()
    if suffix == ".csv":
        save_csv(records, headers, path)
        return
    if suffix in {".xlsx", ".xls"}:
        save_xlsx(records, headers, path)
        return
    raise ValueError(f"不支持的输出类型: {path}")


def collect_headers(records: list[dict]) -> list[str]:
    headers: list[str] = []
    for record in records:
        for key in record.keys():
            if key not in headers:
                headers.append(key)
    return headers


def drop_columns(records: list[dict], columns: list[str]) -> list[dict]:
    drop_set = set(columns)
    return [{k: v for k, v in record.items() if k not in drop_set} for record in records]


def keep_columns(records: list[dict], columns: list[str], required: list[str]) -> list[dict]:
    if not columns:
        return records
    keep_set = set(columns) | set(required)
    return [{k: v for k, v in record.items() if k in keep_set} for record in records]


def normalize_keys(records: list[dict], keys: list[str]) -> list[dict]:
    output: list[dict] = []
    for record in records:
        item = dict(record)
        for key in keys:
            if key not in item:
                raise KeyError(f"缺少连接键: {key}")
            value = item[key]
            if key == "year":
                item[key] = None if value in (None, "") else int(float(value))
            else:
                item[key] = str(value).strip()
        output.append(item)
    return output


def rename_overlaps(
    left_records: list[dict], right_records: list[dict], on: list[str], suffixes: list[str]
) -> tuple[list[dict], list[dict]]:
    left_cols = set().union(*(r.keys() for r in left_records)) if left_records else set()
    right_cols = set().union(*(r.keys() for r in right_records)) if right_records else set()
    overlap = (left_cols & right_cols) - set(on)
    if not overlap:
        return left_records, right_records

    left_suffix, right_suffix = suffixes
    new_left: list[dict] = []
    new_right: list[dict] = []

    for record in left_records:
        item = {}
        for key, value in record.items():
            item[f"{key}{left_suffix}" if key in overlap else key] = value
        new_left.append(item)

    for record in right_records:
        item = {}
        for key, value in record.items():
            item[f"{key}{right_suffix}" if key in overlap else key] = value
        new_right.append(item)

    return new_left, new_right


def make_key(record: dict, on: list[str]) -> tuple:
    return tuple(record.get(k) for k in on)


def merge_records(
    left_records: list[dict],
    right_records: list[dict],
    on: list[str],
    how: str,
    all_headers: list[str],
) -> list[dict]:
    left_by_key: dict[tuple, list[dict]] = {}
    right_by_key: dict[tuple, list[dict]] = {}

    for record in left_records:
        left_by_key.setdefault(make_key(record, on), []).append(record)
    for record in right_records:
        right_by_key.setdefault(make_key(record, on), []).append(record)

    left_keys = set(left_by_key)
    right_keys = set(right_by_key)
    if how == "inner":
        keys = left_keys & right_keys
    elif how == "left":
        keys = left_keys
    elif how == "right":
        keys = right_keys
    else:
        keys = left_keys | right_keys

    merged: list[dict] = []
    for key in sorted(keys):
        left_group = left_by_key.get(key, [None])
        right_group = right_by_key.get(key, [None])
        for left in left_group:
            for right in right_group:
                row = {}
                if left:
                    row.update(left)
                if right:
                    row.update(right)
                for i, col in enumerate(on):
                    row[col] = key[i]
                for header in all_headers:
                    row.setdefault(header, None)
                merged.append(row)
    return merged


def sort_records(records: list[dict], on: list[str]) -> list[dict]:
    def key_func(record: dict) -> tuple:
        return tuple((record.get(col) is None, record.get(col)) for col in on)

    return sorted(records, key=key_func)


def resolve_input_path(raw_path: str) -> Path:
    path = Path(raw_path)
    if path.exists():
        return path

    if not path.is_absolute():
        data_path = Path("data") / path
        if data_path.exists():
            return data_path

    raise FileNotFoundError(f"Input file not found: {raw_path}")


def main() -> None:
    args = parse_args()
    left_path = resolve_input_path(args.left)
    right_path = resolve_input_path(args.right)
    output_path = Path(args.output)

    left_records = read_table(left_path)
    right_records = read_table(right_path)

    left_records = drop_columns(left_records, args.exclude_left)
    right_records = keep_columns(right_records, args.include_right, args.on)

    left_records = normalize_keys(left_records, args.on)
    right_records = normalize_keys(right_records, args.on)

    left_records, right_records = rename_overlaps(
        left_records, right_records, args.on, args.suffixes
    )
    all_headers = collect_headers(left_records) + [
        h for h in collect_headers(right_records) if h not in collect_headers(left_records)
    ]
    merged = merge_records(left_records, right_records, args.on, args.how, all_headers)
    merged = sort_records(merged, args.on)

    save_table(merged, output_path)

    print(f"左表: {left_path}，行数: {len(left_records)}")
    print(f"右表: {right_path}，行数: {len(right_records)}")
    print(f"输出: {output_path}，行数: {len(merged)}")


if __name__ == "__main__":
    main()
