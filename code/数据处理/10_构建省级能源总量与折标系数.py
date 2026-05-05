from __future__ import annotations

from dataclasses import dataclass, asdict
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook


STANDARD_COAL_MJ_PER_KG = 29.3076

ROOT = Path(__file__).resolve().parent
CEADS_DIR = ROOT / "data" / "各省能源清单CEADs"
FACTOR_OUTPUT = ROOT / "data" / "ceads_standard_coal_factors.csv"
PANEL_OUTPUT = ROOT / "data" / "ceads_energy_total_es_2015_2022.csv"


PROVINCE_MAP = {
    "Shanghai": "上海",
    "Yunnan": "云南",
    "InnerMongolia": "内蒙古",
    "Beijing": "北京",
    "Jilin": "吉林",
    "Sichuan": "四川",
    "Tianjin": "天津",
    "Ningxia": "宁夏",
    "Anhui": "安徽",
    "Shandong": "山东",
    "Shanxi": "山西",
    "Guangdong": "广东",
    "Guangxi": "广西",
    "Xinjiang": "新疆",
    "Jiangsu": "江苏",
    "Jiangxi": "江西",
    "Hebei": "河北",
    "Henan": "河南",
    "Zhejiang": "浙江",
    "Hainan": "海南",
    "Hubei": "湖北",
    "Hunan": "湖南",
    "Gansu": "甘肃",
    "Fujian": "福建",
    "Guizhou": "贵州",
    "Liaoning": "辽宁",
    "Chongqing": "重庆",
    "Shaanxi": "陕西",
    "Qinghai": "青海",
    "Heilongjiang": "黑龙江",
}


@dataclass(frozen=True)
class FactorSpec:
    ceads_field: str
    chinese_name: str
    input_unit: str
    factor_to_tce: float
    factor_unit: str
    source_type: str
    source_detail: str
    note: str
    include_in_coal_total: int


def ncv_to_factor_for_ton(ncv_pj_per_1e4_ton: float) -> float:
    mj_per_kg = ncv_pj_per_1e4_ton * 100.0
    return mj_per_kg / STANDARD_COAL_MJ_PER_KG


def ncv_to_factor_for_cubic_meter(ncv_pj_per_1e8_cum: float) -> float:
    mj_per_cum = ncv_pj_per_1e8_cum * 10.0
    return mj_per_cum / STANDARD_COAL_MJ_PER_KG


FACTOR_SPECS = [
    FactorSpec(
        "Raw_Coal",
        "原煤",
        "10^4 ton",
        0.7143,
        "10^4 tce / 10^4 ton",
        "China Energy Statistical Yearbook 2023",
        "各种能源折标准煤参考系数",
        "年鉴图中有直接系数。",
        1,
    ),
    FactorSpec(
        "Cleaned_Coal",
        "洗精煤",
        "10^4 ton",
        0.9000,
        "10^4 tce / 10^4 ton",
        "China Energy Statistical Yearbook 2023",
        "各种能源折标准煤参考系数",
        "年鉴图中有直接系数。",
        1,
    ),
    FactorSpec(
        "Other_Washed_Coal",
        "其他洗煤",
        "10^4 ton",
        0.2857,
        "10^4 tce / 10^4 ton",
        "China Energy Statistical Yearbook 2023",
        "Middlings 下限系数",
        "按保守下限重做，采用年鉴图中可见相关子项的最低系数。",
        1,
    ),
    FactorSpec(
        "Briquettes",
        "煤制品",
        "10^4 ton",
        0.5000,
        "10^4 tce / 10^4 ton",
        "China Energy Statistical Yearbook 2023",
        "Briquettes 参考系数",
        "按保守下限重做，采用煤制品保守参考系数。",
        1,
    ),
    FactorSpec(
        "Coke",
        "焦炭",
        "10^4 ton",
        0.9714,
        "10^4 tce / 10^4 ton",
        "China Energy Statistical Yearbook 2023",
        "各种能源折标准煤参考系数",
        "年鉴图中有直接系数。",
        1,
    ),
    FactorSpec(
        "Coke_Oven_Gas",
        "焦炉煤气",
        "10^8 cu.m",
        0.5714,
        "10^4 tce / 10^8 cu.m",
        "China Energy Statistical Yearbook 2023",
        "Coke Oven Gas 下限系数",
        "按保守下限重做，采用年鉴图区间下限。",
        1,
    ),
    FactorSpec(
        "Other_Gas",
        "其他煤气",
        "10^8 cu.m",
        0.1786,
        "10^4 tce / 10^8 cu.m",
        "China Energy Statistical Yearbook 2023",
        "By Gas Furnace 下限系数",
        "按保守下限重做，采用年鉴图中其他煤气相关子项的最低系数。",
        1,
    ),
    FactorSpec(
        "Other_Coking_Products",
        "其他焦化产品",
        "10^4 ton",
        1.4286,
        "10^4 tce / 10^4 ton",
        "China Energy Statistical Yearbook 2023",
        "Coal Tar/Benzene 下限系数",
        "按保守下限重做，采用年鉴图中其他焦化产品相关子项的最低系数。",
        1,
    ),
    FactorSpec(
        "Crude_Oil",
        "原油",
        "10^4 ton",
        1.4286,
        "10^4 tce / 10^4 ton",
        "China Energy Statistical Yearbook 2023",
        "各种能源折标准煤参考系数",
        "年鉴图中有直接系数。",
        0,
    ),
    FactorSpec(
        "Gasoline",
        "汽油",
        "10^4 ton",
        1.4714,
        "10^4 tce / 10^4 ton",
        "China Energy Statistical Yearbook 2023",
        "各种能源折标准煤参考系数",
        "年鉴图中有直接系数。",
        0,
    ),
    FactorSpec(
        "Kerosene",
        "煤油",
        "10^4 ton",
        1.4714,
        "10^4 tce / 10^4 ton",
        "China Energy Statistical Yearbook 2023",
        "各种能源折标准煤参考系数",
        "年鉴图中有直接系数。",
        0,
    ),
    FactorSpec(
        "Diesel_Oil",
        "柴油",
        "10^4 ton",
        1.4571,
        "10^4 tce / 10^4 ton",
        "China Energy Statistical Yearbook 2023",
        "各种能源折标准煤参考系数",
        "年鉴图中有直接系数。",
        0,
    ),
    FactorSpec(
        "Fuel_Oil",
        "燃料油",
        "10^4 ton",
        1.4286,
        "10^4 tce / 10^4 ton",
        "China Energy Statistical Yearbook 2023",
        "各种能源折标准煤参考系数",
        "年鉴图中有直接系数。",
        0,
    ),
    FactorSpec(
        "LPG",
        "液化石油气",
        "10^4 ton",
        1.7143,
        "10^4 tce / 10^4 ton",
        "China Energy Statistical Yearbook 2023",
        "各种能源折标准煤参考系数",
        "年鉴图中有直接系数。",
        0,
    ),
    FactorSpec(
        "Refinery_Gas",
        "炼厂干气",
        "10^4 ton",
        1.5714,
        "10^4 tce / 10^4 ton",
        "China Energy Statistical Yearbook 2023",
        "各种能源折标准煤参考系数",
        "年鉴图中有直接系数。",
        0,
    ),
    FactorSpec(
        "Other_Petroleum_Products",
        "其他石油制品",
        "10^4 ton",
        1.4286,
        "10^4 tce / 10^4 ton",
        "China Energy Statistical Yearbook 2023",
        "Crude Oil/Fuel Oil 下限系数",
        "按保守下限重做，采用石油类相关已知系数中的保守下限。",
        0,
    ),
    FactorSpec(
        "Natural_Gas",
        "天然气",
        "10^8 cu.m",
        1.1000,
        "10^4 tce / 10^8 cu.m",
        "China Energy Statistical Yearbook 2023",
        "Natural Gas 下限系数",
        "按保守下限重做，采用年鉴图区间下限。",
        0,
    ),
    FactorSpec(
        "Heat",
        "热力",
        "10^10 kJ",
        0.03412,
        "10^4 tce / 10^10 kJ",
        "China Energy Statistical Yearbook 2023",
        "各种能源折标准煤参考系数",
        "年鉴图中有直接系数。",
        0,
    ),
    FactorSpec(
        "Electricity",
        "电力",
        "10^8 kwh",
        0.1229,
        "10^4 tce / 10^8 kwh",
        "China Energy Statistical Yearbook 2023",
        "各种能源折标准煤参考系数",
        "年鉴图中有直接系数。",
        0,
    ),
    FactorSpec(
        "Other_Energy",
        "其他能源",
        "10^4 tce",
        1.0,
        "10^4 tce / 10^4 tce",
        "CEADs workbook unit row",
        "Other_Energy unit = 10^4 tce",
        "原表已是标准煤单位，直接并入能源消费总量。",
        0,
    ),
]


def build_factor_table() -> pd.DataFrame:
    df = pd.DataFrame(asdict(spec) for spec in FACTOR_SPECS)
    df["factor_to_tce"] = df["factor_to_tce"].round(6)
    return df


def extract_year(path: Path) -> int:
    match = re.search(r"(\d{4})", path.stem)
    if not match:
        raise ValueError(f"Cannot parse year from {path.name}")
    return int(match.group(1))


def sheet_to_province(sheet_name: str) -> str:
    province_en = re.sub(r"\d{4}$", "", sheet_name)
    if province_en not in PROVINCE_MAP:
        raise KeyError(f"Unknown province sheet name: {sheet_name}")
    return PROVINCE_MAP[province_en]


def build_panel_table(factor_df: pd.DataFrame) -> pd.DataFrame:
    factor_map = dict(zip(factor_df["ceads_field"], factor_df["factor_to_tce"]))
    coal_fields = factor_df.loc[factor_df["include_in_coal_total"] == 1, "ceads_field"].tolist()

    records = []
    for workbook_path in sorted(CEADS_DIR.glob("*.xlsx")):
        year = extract_year(workbook_path)
        if not 2015 <= year <= 2022:
            continue

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            if sheet_name == "NOTE":
                continue

            ws = workbook[sheet_name]
            rows = ws.iter_rows(values_only=True, min_row=1, max_row=3)
            header = [str(v).strip() if v is not None else "" for v in next(rows)]
            _units = next(rows)
            total_row = next(rows)

            values = {}
            for idx, field in enumerate(header[1:], start=1):
                cell_value = total_row[idx] if idx < len(total_row) else 0
                values[field] = float(cell_value or 0)

            energy_total = sum(values[field] * factor_map[field] for field in factor_map)
            coal_total = sum(values[field] * factor_map[field] for field in coal_fields)
            es = coal_total / energy_total if energy_total else None

            records.append(
                {
                    "year": year,
                    "province": sheet_to_province(sheet_name),
                    "energy_total": round(energy_total, 6),
                    "es": round(es, 6) if es is not None else None,
                }
            )

    panel_df = pd.DataFrame(records).sort_values(["year", "province"]).reset_index(drop=True)
    return panel_df


def main() -> None:
    factor_df = build_factor_table()
    factor_df.to_csv(FACTOR_OUTPUT, index=False, encoding="utf-8-sig")

    panel_df = build_panel_table(factor_df)
    panel_df.to_csv(PANEL_OUTPUT, index=False, encoding="utf-8-sig")

    print(f"Wrote {FACTOR_OUTPUT}")
    print(f"Wrote {PANEL_OUTPUT}")
    print(panel_df.head().to_string(index=False))


if __name__ == "__main__":
    main()
